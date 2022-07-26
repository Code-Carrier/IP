from router import Router
from CSV_processing import csv_processing
import openpyxl


def main(routers,sheet1):

    __rnum__ = len(routers)
    #建立路由器结点数组
    lrouter = []
    for i in range(__rnum__):
        lrouter.append(Router(i,routers[i]))
        lrouter[i].set_dic()

    #交换邻居信息，交换完成后则计算优先级
    for i in range(__rnum__):
        temp_list=lrouter[i].send_message_to_all_neb(0)
        for x in routers[i]:
            lrouter[x].recieve_message(0,i,temp_list)


    #交换优先级信息，设置等待队列
    for i in range(__rnum__):
        temp_list = lrouter[i].send_message_to_all_neb(1)
        for x in routers[i]:
            lrouter[x].recieve_message(1,i,temp_list)
    

    case = "yes"

    #case = input("选择使用的方法--1:只分配一个色块,剩余用随机分配. 2:分配四个色块. 回车键退出")
    case = 1
    while(case < 3):
        sheet_line = 2
        base_col = pow(2,case) + 1
        sheet1.cell(sheet_line,base_col - 1).value = case
        sheet_line += 1
        case = int(case)
        __colornum__ = 4
        if (case != 1) & (case != 2):
            break
        wait_for_update_from_their_neb = []
        right = [0 for i in range(__rnum__)]
        nowtime = 0
        __oldcom__ = 0
        __communication__ = 0
        #每轮所有节点都应当判断自己本轮是否可以参与分配色块
        go = True
        correct = 0
        while go:
            nowtime = nowtime + 1
            wait_for_update_because_finish = []
            for i in range(__rnum__):
                if right[i] < __colornum__:
                    ok,count = lrouter[i].select_turn(case)
                else:
                    continue
                if ok:
                    wait_for_update_from_their_neb.append(i)
                    right[i] += count
                    if right[i] >= __colornum__:
                        correct = correct + 1
                        wait_for_update_because_finish.append(i)

            #每轮分配完成后应当更新所有节点的等待队列
            for elem in wait_for_update_from_their_neb:
                temp_list = lrouter[elem].send_message_to_all_neb(2)
                for x in routers[elem]:
                    ok = lrouter[x].recieve_message(2,elem,temp_list)
                    __communication__ += 1
                    if not ok:
                        return "有冲突"
            wait_for_update_from_their_neb.clear()
            
            for elem in wait_for_update_because_finish:
                for x in routers[elem]:
                    lrouter[x].remove_neb_who_finished(elem)
            wait_for_update_because_finish.clear()
            print("当前交换邻居信息，当前阶段共通信 ",__communication__-__oldcom__," 次")
            print("一共通信 ",__communication__," 次")
            __oldcom__ = __communication__
            #如果所有节点都分配结束，应当结束循环
            if correct == __rnum__:
                go = False
        
        #输出本次分配所有节点的第一个色块号中的最大块号
        temp_list = []
        for i in range(__rnum__):
            temp = lrouter[i].send_choose()   
            if temp[0] not in temp_list:
                temp_list.append(temp[0])
        
        print("每个节点的第一个色块共需要 ",len(temp_list)," 个颜色。")
        print("总共交互轮次为 ",nowtime)
        for i in range(__rnum__):
            temp_list = lrouter[i].send_choose()
            sheet1.cell(sheet_line,1).value = i
            for j in range(len(temp_list)):
                sheet1.cell(sheet_line,j+base_col).value = temp_list[j]
            sheet_line += 1
            lrouter[i].init_again()
        nowtime = 0
        __oldcom__ = 0
        __communication__ = 0
        #case = input("选择使用的方法--1:只分配一个色块,剩余用随机分配. 2:分配四个色块. 回车键退出")
        case += 1


if __name__ == "__main__":
    name = input("输入文件名：")
    f = openpyxl.Workbook()
    name_list = ["S","G","H"]
    #for name in name_list:
    while(name):
        sheet1 = f.create_sheet(name) #创建sheet
        name += "_total"
        routers = csv_processing(name)
        ok = main(routers,sheet1)
        print(ok)
        name = input("输入文件名：")
    f.save("./output/"+"output.xlsx")
    
    